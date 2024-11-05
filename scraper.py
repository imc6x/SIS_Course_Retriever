import requests
import openpyxl
import pandas as pd
import os
from time import sleep

def create_excel_sheet(filename):

    """Creates Excel sheet with the name UVA_SIS_Course_Retriever if it does not already exist
    """
    if not os.path.exists(filename):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet['A1'] = 'Subject'
        sheet['B1'] = 'Course Number'
        sheet['C1'] = 'Instructor'
        workbook.save(filename)
        print(f"Created new Excel file: {filename}")
    else:
        print(f"File {filename} already exists")

    
    
    
def get_courses(filename):
    
    seen_courses_and_instructors = set()
    
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    department_url = "https://sisuva.admin.virginia.edu/psc/ihprd/UVSS/SA/s/WEBLIB_HCX_CM.H_CLASS_SEARCH.FieldFormula.IScript_ClassSearchOptions"
    department_response = requests.get(department_url, params={
        "institution": "UVA01", 
        "term": "1248"
    })
    departments = department_response.json()['subjects']
    
    row = 2
    
    for dept in departments:
        subject = dept['subject']
        print(f"Getting courses for {subject}")
        
        page = 1
        while True:
            course_url = "https://sisuva.admin.virginia.edu/psc/ihprd/UVSS/SA/s/WEBLIB_HCX_CM.H_CLASS_SEARCH.FieldFormula.IScript_ClassSearch"
            params = {
                "institution": "UVA01",
                "term": "1248",
                "subject": subject,
                "page": str(page)
            }
            
            courses = requests.get(course_url, params=params).json()
        
            if not courses:
                break
            
            for course in courses:
                sheet.cell(row=row, column=1, value=subject)
                sheet.cell(row=row, column=2, value=course['catalog_nbr'])
                row += 1
            
            page += 1
            sleep(0.5)  
        
        workbook.save(filename)


if __name__ == "__main__":
    file = 'UVA_SIS_Course_Retriever.xlsx'
    create_excel_sheet(file)
    get_courses(file)


    
    
    
    
